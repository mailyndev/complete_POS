from fastapi import APIRouter, Depends, HTTPException, Query, Request
from sqlmodel import Session, select, or_
from typing import List, Optional, Dict, Any
from datetime import datetime, date, timedelta

from ..database.connection import get_session
from ..database.schema import (
    Producto, Categoria, Subcategoria, Marca, Impuesto, Inventario,
    MovimientoInventario, Lote, HistorialPrecios, HistorialCostos, Usuario, Auditoria,
    Transferencia, DetalleTransferencia, Sucursal, Empresa, Cabys
)
from ..utils.security import get_current_user, PermissionChecker
import base64
import csv
import io
import openpyxl
import os

router = APIRouter(prefix="/inventory", tags=["Inventario"])

# --- PRODUCTOS CRUD ---

@router.get("/products", response_model=List[Dict[str, Any]])
def get_products(
    search: Optional[str] = Query(None, description="Buscar por SKU, código o nombre"),
    categoria_id: Optional[int] = Query(None),
    stock_bajo: Optional[bool] = Query(None),
    include_inactive: Optional[bool] = Query(False, description="Incluir productos inactivos"),
    current_user: Usuario = Depends(get_current_user),
    session: Session = Depends(get_session)
):
    # Trae productos y adjunta información de impuestos, stock por sucursal y categoría
    query = select(Producto)
    if not include_inactive:
        query = query.where(Producto.activo == True)
    
    if search:
        query = query.where(
            or_(
                Producto.sku.like(f"%{search}%"),
                Producto.codigo_barras.like(f"%{search}%"),
                Producto.nombre.like(f"%{search}%")
            )
        )
    
    if categoria_id:
        query = query.join(Subcategoria).where(Subcategoria.categoria_id == categoria_id)
        
    products = session.exec(query).all()
    results = []
    
    for p in products:
        # Calcular stock total e inventario de la sucursal del usuario
        stmt_inv = select(Inventario).where(
            Inventario.producto_id == p.id,
            Inventario.sucursal_id == current_user.sucursal_id
        )
        inv = session.exec(stmt_inv).first()
        existencia = inv.existencia_actual if inv else 0.0
        
        # Filtro de stock bajo
        if stock_bajo and existencia > p.stock_minimo:
            continue
            
        # Calcular margen
        utilidad = p.precio_venta - p.precio_costo
        margen = (utilidad / p.precio_venta * 100) if p.precio_venta > 0 else 0.0
        
        results.append({
            "id": p.id,
            "sku": p.sku,
            "codigo_barras": p.codigo_barras,
            "nombre": p.nombre,
            "descripcion": p.descripcion,
            "marca": p.marca.nombre,
            "marca_id": p.marca_id,
            "subcategoria": p.subcategoria.nombre,
            "subcategoria_id": p.subcategoria_id,
            "categoria": p.subcategoria.categoria.nombre,
            "categoria_id": p.subcategoria.categoria_id,
            "unidad_medida": p.unidad_medida,
            "precio_costo": p.precio_costo,
            "precio_venta": p.precio_venta,
            "precio_mayorista": p.precio_mayorista,
            "impuesto": p.impuesto.nombre,
            "impuesto_porcentaje": p.impuesto.porcentaje,
            "impuesto_id": p.impuesto_id,
            "stock_minimo": p.stock_minimo,
            "stock_maximo": p.stock_maximo,
            "existencia": existencia,
            "imagen_path": p.imagen_path,
            "activo": p.activo,
            "proveedor_id": p.proveedor_id,
            "codigo_cabys": p.codigo_cabys,
            "utilidad": round(utilidad, 2),
            "margen_bruto": round(margen, 2)
        })
        
    return results

@router.post("/products", status_code=201)
def create_product(
    payload: Dict[str, Any],
    request: Request,
    current_user: Usuario = Depends(PermissionChecker("inventory:edit")),
    session: Session = Depends(get_session)
):
    sku = payload.get("sku")
    codigo_barras = payload.get("codigo_barras")
    nombre = payload.get("nombre")
    marca_id = payload.get("marca_id")
    subcategoria_id = payload.get("subcategoria_id")
    impuesto_id = payload.get("impuesto_id")
    precio_costo = float(payload.get("precio_costo", 0))
    precio_venta = float(payload.get("precio_venta", 0))
    precio_mayorista = float(payload.get("precio_mayorista", 0))
    stock_minimo = float(payload.get("stock_minimo", 0))
    stock_maximo = float(payload.get("stock_maximo", 0))
    unidad_medida = payload.get("unidad_medida", "Unidad")
    descripcion = payload.get("descripcion", "")
    imagen_path = payload.get("imagen_path")
    codigo_cabys = payload.get("codigo_cabys")
    if codigo_cabys == "" or codigo_cabys is None:
        codigo_cabys = None
    
    proveedor_id = payload.get("proveedor_id")
    if proveedor_id == "" or proveedor_id is None:
        proveedor_id = None
    else:
        proveedor_id = int(proveedor_id)
        
    existencia_inicial = float(payload.get("existencia_inicial", 0.0))
    
    # Validaciones
    if not sku or not codigo_barras or not nombre or not marca_id or not subcategoria_id or not impuesto_id:
        raise HTTPException(status_code=400, detail="Faltan campos obligatorios para crear el producto")
        
    if precio_costo < 0 or precio_venta < 0 or precio_mayorista < 0:
        raise HTTPException(status_code=400, detail="Los precios y costos no pueden ser negativos")
        
    if stock_minimo < 0 or stock_maximo < 0:
        raise HTTPException(status_code=400, detail="El stock mínimo y máximo no pueden ser negativos")
        
    if stock_maximo < stock_minimo:
        raise HTTPException(status_code=400, detail="El stock máximo no puede ser menor al stock mínimo")
        
    if existencia_inicial < 0:
        raise HTTPException(status_code=400, detail="La existencia inicial no puede ser negativa")

    # Validar duplicados
    dup = session.exec(select(Producto).where(or_(Producto.sku == sku, Producto.codigo_barras == codigo_barras))).first()
    if dup:
        raise HTTPException(status_code=400, detail="Ya existe un producto con el mismo SKU o código de barras")

    producto = Producto(
        sku=sku,
        codigo_barras=codigo_barras,
        nombre=nombre,
        descripcion=descripcion,
        marca_id=marca_id,
        subcategoria_id=subcategoria_id,
        unidad_medida=unidad_medida,
        precio_costo=precio_costo,
        precio_venta=precio_venta,
        precio_mayorista=precio_mayorista,
        impuesto_id=impuesto_id,
        stock_minimo=stock_minimo,
        stock_maximo=stock_maximo,
        imagen_path=imagen_path,
        proveedor_id=proveedor_id,
        codigo_cabys=codigo_cabys,
        activo=True
    )
    session.add(producto)
    session.commit()
    session.refresh(producto)

    # Inicializar inventario en la sucursal actual
    inventario = Inventario(
        sucursal_id=current_user.sucursal_id,
        producto_id=producto.id,
        existencia_actual=existencia_inicial
    )
    session.add(inventario)
    session.commit()
    session.refresh(inventario)

    # Registrar lote y movimiento si hay stock inicial
    if existencia_inicial > 0.0:
        vencimiento = date.today() + timedelta(days=365)
        lote = Lote(
            producto_id=producto.id,
            numero_lote="LOT-INICIAL",
            fecha_ingreso=date.today(),
            fecha_vencimiento=vencimiento,
            stock_inicial=existencia_inicial,
            stock_actual=existencia_inicial,
            costo_unitario=precio_costo
        )
        session.add(lote)
        
        mov = MovimientoInventario(
            inventario_id=inventario.id,
            tipo_movimiento="entrada",
            cantidad=existencia_inicial,
            motivo="Inventario Inicial",
            usuario_id=current_user.id
        )
        session.add(mov)
    
    # Registrar auditoría
    audit = Auditoria(
        usuario_id=current_user.id,
        accion="CREAR_PRODUCTO",
        modulo="Inventario",
        detalles=f"Producto creado: {producto.nombre} ({producto.sku}). Existencia inicial: {existencia_inicial}. Precio: ₡{producto.precio_venta}",
        ip_address=request.client.host
    )
    session.add(audit)
    session.commit()

    return {"message": "Producto creado con éxito", "id": producto.id}

@router.put("/products/{id}")
def update_product(
    id: int,
    payload: Dict[str, Any],
    request: Request,
    current_user: Usuario = Depends(PermissionChecker("inventory:edit")),
    session: Session = Depends(get_session)
):
    producto = session.get(Producto, id)
    if not producto:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
        
    precio_anterior = producto.precio_venta
    costo_anterior = producto.precio_costo
    
    # Actualizar campos
    producto.nombre = payload.get("nombre", producto.nombre)
    producto.descripcion = payload.get("descripcion", producto.descripcion)
    
    new_sku = payload.get("sku", producto.sku)
    new_barcode = payload.get("codigo_barras", producto.codigo_barras)
    if new_sku != producto.sku or new_barcode != producto.codigo_barras:
        dup = session.exec(select(Producto).where(
            (Producto.id != producto.id) & 
            ((Producto.sku == new_sku) | (Producto.codigo_barras == new_barcode))
        )).first()
        if dup:
            raise HTTPException(status_code=400, detail="Ya existe otro producto con el mismo SKU o código de barras")
            
    producto.sku = new_sku
    producto.codigo_barras = new_barcode
    producto.marca_id = payload.get("marca_id", producto.marca_id)
    producto.subcategoria_id = payload.get("subcategoria_id", producto.subcategoria_id)
    producto.unidad_medida = payload.get("unidad_medida", producto.unidad_medida)
    producto.stock_minimo = float(payload.get("stock_minimo", producto.stock_minimo))
    producto.stock_maximo = float(payload.get("stock_maximo", producto.stock_maximo))
    producto.imagen_path = payload.get("imagen_path", producto.imagen_path)
    producto.impuesto_id = payload.get("impuesto_id", producto.impuesto_id)
    
    if "activo" in payload:
        act_val = payload.get("activo")
        if isinstance(act_val, str):
            producto.activo = act_val.lower() == "true"
        else:
            producto.activo = bool(act_val)
            
    if "proveedor_id" in payload:
        prov_val = payload.get("proveedor_id")
        if prov_val == "" or prov_val is None:
            producto.proveedor_id = None
        else:
            producto.proveedor_id = int(prov_val)
            
    if "codigo_cabys" in payload:
        cabys_val = payload.get("codigo_cabys")
        if cabys_val == "" or cabys_val is None:
            producto.codigo_cabys = None
        else:
            producto.codigo_cabys = str(cabys_val).strip()
            
    precio_nuevo = float(payload.get("precio_venta", producto.precio_venta))
    costo_nuevo = float(payload.get("precio_costo", producto.precio_costo))
    producto.precio_mayorista = float(payload.get("precio_mayorista", producto.precio_mayorista))

    # Validaciones
    if costo_nuevo < 0 or precio_nuevo < 0 or producto.precio_mayorista < 0:
        raise HTTPException(status_code=400, detail="Los precios y costos no pueden ser negativos")
        
    if producto.stock_minimo < 0 or producto.stock_maximo < 0:
        raise HTTPException(status_code=400, detail="El stock mínimo y máximo no pueden ser negativos")
        
    if producto.stock_maximo < producto.stock_minimo:
        raise HTTPException(status_code=400, detail="El stock máximo no puede ser menor al stock mínimo")
    
    # Historial de cambios de precio de venta
    if precio_nuevo != precio_anterior:
        hist_p = HistorialPrecios(
            producto_id=producto.id,
            precio_anterior=precio_anterior,
            precio_nuevo=precio_nuevo,
            motivo=payload.get("motivo_cambio_precio", "Actualización administrativa de catálogo"),
            usuario_id=current_user.id
        )
        session.add(hist_p)
        producto.precio_venta = precio_nuevo
        
    # Historial de costos
    if costo_nuevo != costo_anterior:
        hist_c = HistorialCostos(
            producto_id=producto.id,
            costo_anterior=costo_anterior,
            costo_nuevo=costo_nuevo,
            motivo=payload.get("motivo_cambio_costo", "Actualización administrativa de catálogo"),
            usuario_id=current_user.id
        )
        session.add(hist_c)
        producto.precio_costo = costo_nuevo
        
    session.add(producto)
    
    # Registrar auditoría
    audit = Auditoria(
        usuario_id=current_user.id,
        accion="MODIFICAR_PRODUCTO",
        modulo="Inventario",
        detalles=f"Producto modificado: {producto.nombre} ({producto.sku}).",
        ip_address=request.client.host
    )
    session.add(audit)
    session.commit()
    
    return {"message": "Producto actualizado con éxito"}

# --- LOTES Y ALERTAS DE EXPIRACIÓN ---

@router.get("/lots", response_model=List[Dict[str, Any]])
def get_lots(
    producto_id: Optional[int] = Query(None),
    current_user: Usuario = Depends(get_current_user),
    session: Session = Depends(get_session)
):
    query = select(Lote)
    if producto_id:
        query = query.where(Lote.producto_id == producto_id)
        
    lots = session.exec(query).all()
    results = []
    
    hoy = date.today()
    
    for l in lots:
        # Calcular nivel de alerta basado en la categoría
        p = l.producto
        dias_config = p.subcategoria.categoria.dias_alerta_vencimiento
        dias_restantes = (l.fecha_vencimiento - hoy).days
        
        # Informativa, Advertencia, Crítica
        if dias_restantes <= 3:
            alerta = "Crítica"
        elif dias_restantes <= dias_config:
            alerta = "Advertencia"
        elif dias_restantes <= (dias_config + 15):
            alerta = "Informativa"
        else:
            alerta = "Ninguna"
            
        results.append({
            "id": l.id,
            "producto_id": l.producto_id,
            "producto_nombre": p.nombre,
            "numero_lote": l.numero_lote,
            "fecha_ingreso": l.fecha_ingreso,
            "fecha_vencimiento": l.fecha_vencimiento,
            "stock_inicial": l.stock_inicial,
            "stock_actual": l.stock_actual,
            "dias_restantes": dias_restantes,
            "alerta": alerta
        })
        
    return results

@router.post("/lots", status_code=201)
def create_lot(
    payload: Dict[str, Any],
    request: Request,
    current_user: Usuario = Depends(PermissionChecker("inventory:edit")),
    session: Session = Depends(get_session)
):
    producto_id = payload.get("producto_id")
    numero_lote = payload.get("numero_lote")
    fecha_vencimiento_str = payload.get("fecha_vencimiento")
    stock_inicial = float(payload.get("stock_inicial", 0))
    
    if not producto_id or not numero_lote or not fecha_vencimiento_str:
        raise HTTPException(status_code=400, detail="Faltan datos obligatorios para el lote")
        
    try:
        fecha_vencimiento = datetime.strptime(fecha_vencimiento_str, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido. Usar YYYY-MM-DD")
        
    # Crear Lote
    lote = Lote(
        producto_id=producto_id,
        numero_lote=numero_lote,
        fecha_vencimiento=fecha_vencimiento,
        stock_inicial=stock_inicial,
        stock_actual=stock_inicial
    )
    session.add(lote)
    
    # Aumentar existencia en inventario
    stmt_inv = select(Inventario).where(
        Inventario.producto_id == producto_id,
        Inventario.sucursal_id == current_user.sucursal_id
    )
    inv = session.exec(stmt_inv).first()
    if not inv:
        inv = Inventario(sucursal_id=current_user.sucursal_id, producto_id=producto_id, existencia_actual=0.0)
    
    inv.existencia_actual += stock_inicial
    session.add(inv)
    session.commit()
    session.refresh(inv)
    
    # Crear movimiento de inventario
    mov = MovimientoInventario(
        inventario_id=inv.id,
        tipo_movimiento="entrada",
        cantidad=stock_inicial,
        motivo=f"Entrada por ingreso de lote: {numero_lote}",
        usuario_id=current_user.id
    )
    session.add(mov)
    
    # Auditoría
    audit = Auditoria(
        usuario_id=current_user.id,
        accion="CREAR_LOTE",
        modulo="Inventario",
        detalles=f"Ingreso de lote {numero_lote} para producto {lote.producto.nombre} con stock de {stock_inicial} unidades.",
        ip_address=request.client.host
    )
    session.add(audit)
    session.commit()
    
    return {"message": "Lote ingresado con éxito"}

# --- AUXILIARES ---

@router.get("/categories")
def get_categories(session: Session = Depends(get_session)):
    categories = session.exec(select(Categoria)).all()
    results = []
    for c in categories:
        subcats = session.exec(select(Subcategoria).where(Subcategoria.categoria_id == c.id)).all()
        results.append({
            "id": c.id,
            "nombre": c.nombre,
            "dias_alerta_vencimiento": c.dias_alerta_vencimiento,
            "subcategorias": [{"id": s.id, "nombre": s.nombre} for s in subcats]
        })
    return results

@router.get("/marcas")
def get_marcas(session: Session = Depends(get_session)):
    return session.exec(select(Marca)).all()

@router.get("/taxes")
def get_taxes(session: Session = Depends(get_session)):
    return session.exec(select(Impuesto).where(Impuesto.activo == True)).all()

@router.get("/kardex/{product_id}")
def get_kardex(product_id: int, current_user: Usuario = Depends(get_current_user), session: Session = Depends(get_session)):
    stmt = select(MovimientoInventario).join(Inventario).where(
        Inventario.producto_id == product_id,
        Inventario.sucursal_id == current_user.sucursal_id
    ).order_by(MovimientoInventario.fecha_registro.desc())
    
    movs = session.exec(stmt).all()
    results = []
    
    for m in movs:
        # Obtener usuario responsable
        usr = session.get(Usuario, m.usuario_id)
        results.append({
            "id": m.id,
            "fecha": m.fecha_registro,
            "tipo": m.tipo_movimiento, # entrada, salida, ajuste, transferencia
            "cantidad": m.cantidad,
            "motivo": m.motivo,
            "usuario": usr.nombre if usr else "Desconocido"
        })
    return results

@router.post("/adjust", status_code=200)
def adjust_inventory(
    payload: Dict[str, Any],
    request: Request,
    current_user: Usuario = Depends(PermissionChecker("inventory:edit")),
    session: Session = Depends(get_session)
):
    producto_id = payload.get("producto_id")
    tipo_ajuste = payload.get("tipo_ajuste", "fijar") # "aumentar", "disminuir", "fijar"
    cantidad_variacion = float(payload.get("cantidad", 0))
    motivo = payload.get("motivo", "Ajuste manual de inventario")
    
    if not producto_id:
        raise HTTPException(status_code=400, detail="ID del producto es requerido")
        
    inv = session.exec(select(Inventario).where(
        Inventario.producto_id == producto_id,
        Inventario.sucursal_id == current_user.sucursal_id
    )).first()
    
    if not inv:
        inv = Inventario(sucursal_id=current_user.sucursal_id, producto_id=producto_id, existencia_actual=0.0)
        session.add(inv)
        session.commit()
        session.refresh(inv)
        
    existencia_anterior = inv.existencia_actual
    
    if tipo_ajuste == "aumentar":
        cantidad_ajustada = existencia_anterior + cantidad_variacion
    elif tipo_ajuste == "disminuir":
        cantidad_ajustada = existencia_anterior - cantidad_variacion
    elif tipo_ajuste == "fijar":
        cantidad_ajustada = cantidad_variacion
    else:
        raise HTTPException(status_code=400, detail="Tipo de ajuste no válido")
        
    if cantidad_ajustada < 0.0:
        raise HTTPException(status_code=400, detail="El stock resultante no puede ser negativo")
        
    diferencia = cantidad_ajustada - existencia_anterior
    if diferencia == 0:
        return {"message": "El stock no presenta variación."}
        
    inv.existencia_actual = cantidad_ajustada
    session.add(inv)
    
    # Crear movimiento
    tipo = "entrada" if diferencia > 0 else "salida"
    mov = MovimientoInventario(
        inventario_id=inv.id,
        tipo_movimiento=tipo,
        cantidad=abs(diferencia),
        motivo=f"Ajuste manual ({tipo_ajuste}): {motivo}. Ajustado de {existencia_anterior} a {cantidad_ajustada}",
        usuario_id=current_user.id
    )
    session.add(mov)
    
    # Auditoría
    audit = Auditoria(
        usuario_id=current_user.id,
        accion="AJUSTE_INVENTARIO",
        modulo="Inventario",
        detalles=f"Ajuste manual de {inv.producto.nombre} ({tipo_ajuste}). Variación: {diferencia} unidades. Motivo: {motivo}",
        ip_address=request.client.host
    )
    session.add(audit)
    session.commit()
    
    return {"message": "Ajuste de inventario realizado con éxito"}

@router.get("/branches", response_model=List[Dict[str, Any]])
def get_branches(session: Session = Depends(get_session)):
    branches = session.exec(select(Sucursal).where(Sucursal.activa == True)).all()
    return [{"id": b.id, "nombre": b.nombre, "direccion": b.direccion, "telefono": b.telefono} for b in branches]

@router.get("/transfers", response_model=List[Dict[str, Any]])
def get_transfers(current_user: Usuario = Depends(get_current_user), session: Session = Depends(get_session)):
    transfers = session.exec(select(Transferencia).order_by(Transferencia.fecha.desc())).all()
    results = []
    for t in transfers:
        origin = session.get(Sucursal, t.sucursal_origen_id)
        dest = session.get(Sucursal, t.sucursal_destino_id)
        usr = session.get(Usuario, t.usuario_id)
        
        detalles = []
        for d in t.detalles:
            prod = session.get(Producto, d.producto_id)
            detalles.append({
                "id": d.id,
                "producto_id": d.producto_id,
                "producto_nombre": prod.nombre if prod else "Producto Eliminado",
                "sku": prod.sku if prod else "",
                "cantidad": d.cantidad
            })
        
        results.append({
            "id": t.id,
            "sucursal_origen_id": t.sucursal_origen_id,
            "sucursal_origen_nombre": origin.nombre if origin else "Desconocida",
            "sucursal_destino_id": t.sucursal_destino_id,
            "sucursal_destino_nombre": dest.nombre if dest else "Desconocida",
            "usuario_id": t.usuario_id,
            "usuario_nombre": usr.nombre if usr else "Desconocido",
            "fecha": t.fecha,
            "estado": t.estado,
            "detalles": detalles
        })
    return results

@router.post("/transfers", status_code=201)
def create_transfer(
    payload: Dict[str, Any],
    request: Request,
    current_user: Usuario = Depends(PermissionChecker("inventory:edit")),
    session: Session = Depends(get_session)
):
    sucursal_origen_id = payload.get("sucursal_origen_id")
    sucursal_destino_id = payload.get("sucursal_destino_id")
    items = payload.get("items")
    
    if not sucursal_origen_id or not sucursal_destino_id or not items:
        raise HTTPException(status_code=400, detail="Faltan datos obligatorios para la transferencia")
        
    if sucursal_origen_id == sucursal_destino_id:
        raise HTTPException(status_code=400, detail="La sucursal de origen y destino no pueden ser la misma")
        
    # Crear transferencia
    transferencia = Transferencia(
        sucursal_origen_id=sucursal_origen_id,
        sucursal_destino_id=sucursal_destino_id,
        usuario_id=current_user.id,
        estado="solicitada"
    )
    session.add(transferencia)
    session.commit()
    session.refresh(transferencia)
    
    for item in items:
        det = DetalleTransferencia(
            transferencia_id=transferencia.id,
            producto_id=item["producto_id"],
            cantidad=float(item["cantidad"])
        )
        session.add(det)
        
    # Auditoria
    audit = Auditoria(
        usuario_id=current_user.id,
        accion="SOLICITAR_TRANSFERENCIA",
        modulo="Inventario",
        detalles=f"Solicitud de transferencia #{transferencia.id} desde sucursal ID {sucursal_origen_id} a ID {sucursal_destino_id}",
        ip_address=request.client.host
    )
    session.add(audit)
    session.commit()
    
    return {"message": "Transferencia solicitada con éxito", "id": transferencia.id}

@router.post("/transfers/{id}/dispatch")
def dispatch_transfer(
    id: int,
    request: Request,
    current_user: Usuario = Depends(PermissionChecker("inventory:edit")),
    session: Session = Depends(get_session)
):
    transfer = session.get(Transferencia, id)
    if not transfer:
        raise HTTPException(status_code=404, detail="Transferencia no encontrada")
        
    if transfer.estado != "solicitada":
        raise HTTPException(status_code=400, detail=f"No se puede despachar una transferencia en estado: {transfer.estado}")
        
    # Validar existencias en la sucursal de origen
    for d in transfer.detalles:
        inv_origen = session.exec(select(Inventario).where(
            Inventario.producto_id == d.producto_id,
            Inventario.sucursal_id == transfer.sucursal_origen_id
        )).first()
        
        if not inv_origen or inv_origen.existencia_actual < d.cantidad:
            prod = session.get(Producto, d.producto_id)
            nombre_prod = prod.nombre if prod else f"ID {d.producto_id}"
            stock_disp = inv_origen.existencia_actual if inv_origen else 0.0
            raise HTTPException(
                status_code=400, 
                detail=f"Stock insuficiente para {nombre_prod} en sucursal origen. Requerido: {d.cantidad}, Disponible: {stock_disp}"
            )
            
    # Proceder con el descuento de stock de origen y registro de movimientos
    for d in transfer.detalles:
        inv_origen = session.exec(select(Inventario).where(
            Inventario.producto_id == d.producto_id,
            Inventario.sucursal_id == transfer.sucursal_origen_id
        )).first()
        
        inv_origen.existencia_actual -= d.cantidad
        session.add(inv_origen)
        
        # Registrar Kárdex (Salida por transferencia)
        mov = MovimientoInventario(
            inventario_id=inv_origen.id,
            tipo_movimiento="transferencia",
            cantidad=d.cantidad,
            motivo=f"Salida por despacho de transferencia #{transfer.id} hacia sucursal ID {transfer.sucursal_destino_id}",
            usuario_id=current_user.id
        )
        session.add(mov)
        
    transfer.estado = "despachada"
    session.add(transfer)
    
    # Auditoría
    audit = Auditoria(
        usuario_id=current_user.id,
        accion="DESPACHAR_TRANSFERENCIA",
        modulo="Inventario",
        detalles=f"Despacho de transferencia #{transfer.id} realizado. Stock retirado de origen.",
        ip_address=request.client.host
    )
    session.add(audit)
    session.commit()
    
    return {"message": "Transferencia despachada con éxito"}

@router.post("/transfers/{id}/receive")
def receive_transfer(
    id: int,
    request: Request,
    current_user: Usuario = Depends(PermissionChecker("inventory:edit")),
    session: Session = Depends(get_session)
):
    transfer = session.get(Transferencia, id)
    if not transfer:
        raise HTTPException(status_code=404, detail="Transferencia no encontrada")
        
    if transfer.estado != "despachada":
        raise HTTPException(status_code=400, detail=f"No se puede recibir una transferencia en estado: {transfer.estado}")
        
    # Aumentar existencias en la sucursal de destino
    for d in transfer.detalles:
        inv_destino = session.exec(select(Inventario).where(
            Inventario.producto_id == d.producto_id,
            Inventario.sucursal_id == transfer.sucursal_destino_id
        )).first()
        
        if not inv_destino:
            inv_destino = Inventario(
                sucursal_id=transfer.sucursal_destino_id,
                producto_id=d.producto_id,
                existencia_actual=0.0
            )
            session.add(inv_destino)
            session.commit()
            session.refresh(inv_destino)
            
        inv_destino.existencia_actual += d.cantidad
        session.add(inv_destino)
        
        # Registrar Kárdex (Entrada por transferencia)
        mov = MovimientoInventario(
            inventario_id=inv_destino.id,
            tipo_movimiento="transferencia",
            cantidad=d.cantidad,
            motivo=f"Entrada por recepción de transferencia #{transfer.id} desde sucursal ID {transfer.sucursal_origen_id}",
            usuario_id=current_user.id
        )
        session.add(mov)
        
    transfer.estado = "recibida"
    session.add(transfer)
    
    # Auditoría
    audit = Auditoria(
        usuario_id=current_user.id,
        accion="RECIBIR_TRANSFERENCIA",
        modulo="Inventario",
        detalles=f"Recepción de transferencia #{transfer.id} realizada. Stock ingresado a destino.",
        ip_address=request.client.host
    )
    session.add(audit)
    session.commit()
    
    return {"message": "Transferencia recibida con éxito"}

@router.delete("/products/{id}", status_code=200)
def delete_product(
    id: int,
    request: Request,
    current_user: Usuario = Depends(PermissionChecker("inventory:edit")),
    session: Session = Depends(get_session)
):
    producto = session.get(Producto, id)
    if not producto:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
        
    # Validar transacciones asociadas
    from ..database.schema import DetalleVenta, DetalleCompra
    has_sales = session.exec(select(DetalleVenta).where(DetalleVenta.producto_id == id)).first()
    has_purchases = session.exec(select(DetalleCompra).where(DetalleCompra.producto_id == id)).first()
    
    has_history = bool(has_sales or has_purchases)
    
    # Aplicar borrado lógico
    producto.activo = False
    session.add(producto)
    
    detalles_audit = f"Producto desactivado (borrado lógico): {producto.nombre} ({producto.sku})."
    if has_history:
        detalles_audit += " Nota: El producto tiene transacciones históricas, se conservó por integridad referencial."
        
    # Auditoría
    audit = Auditoria(
        usuario_id=current_user.id,
        accion="ELIMINAR_PRODUCTO",
        modulo="Inventario",
        detalles=detalles_audit,
        ip_address=request.client.host
    )
    session.add(audit)
    session.commit()
    
    return {"message": "Producto eliminado con éxito (desactivado)"}

# --- IMPORTACIÓN Y EXPORTACIÓN MASIVA ---

@router.get("/import/template")
def get_import_template():
    from fastapi.responses import StreamingResponse
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "SKU", "Codigo de barras", "Nombre", "Descripcion", 
        "Categoria", "Subcategoria", "Marca", "Unidad de medida", 
        "Precio costo", "Precio venta", "Precio mayorista", "Impuesto"
    ])
    writer.writerow([
        "PROD-001", "744100123456", "Leche Semidescremada 1L", "Leche fluida pasteurizada",
        "Lácteos", "Leche", "Dos Pinos", "Unidad",
        "850.00", "1050.00", "980.00", "IVA 1%"
    ])
    output.seek(0)
    content = output.getvalue()
    return StreamingResponse(
        io.BytesIO(content.encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=plantilla_productos.csv"}
    )

def parse_csv(content_bytes: bytes) -> List[Dict[str, str]]:
    try:
        decoded = content_bytes.decode("utf-8")
    except UnicodeDecodeError:
        decoded = content_bytes.decode("latin-1")
    reader = csv.DictReader(io.StringIO(decoded))
    rows = []
    for row in reader:
        clean_row = {str(k).strip(): str(v).strip() for k, v in row.items() if k is not None}
        rows.append(clean_row)
    return rows

def parse_xlsx(content_bytes: bytes) -> List[Dict[str, str]]:
    wb = openpyxl.load_workbook(io.BytesIO(content_bytes), data_only=True)
    sheet = wb.active
    rows = []
    headers = []
    for col in range(1, sheet.max_column + 1):
        val = sheet.cell(row=1, column=col).value
        headers.append(str(val).strip() if val is not None else f"Column_{col}")
    for r in range(2, sheet.max_row + 1):
        row_cells = [sheet.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
        if all(x is None for x in row_cells):
            continue
        row_dict = {}
        for c, h in enumerate(headers):
            val = sheet.cell(row=r, column=c+1).value
            row_dict[h] = str(val).strip() if val is not None else ""
        rows.append(row_dict)
    return rows

def map_fields(row: Dict[str, str]) -> Dict[str, Any]:
    mapping = {
        "sku": ["sku", "SKU"],
        "codigo_barras": ["codigo_barras", "codigo de barras", "código de barras", "barras", "barcode", "Código de barras"],
        "nombre": ["nombre", "nombre producto", "producto", "Nombre"],
        "descripcion": ["descripcion", "descripción", "detalle", "Descripción"],
        "categoria": ["categoria", "categoría", "Categoría"],
        "subcategoria": ["subcategoria", "subcategoría", "Subcategoría"],
        "marca": ["marca", "Marca"],
        "unidad_medida": ["unidad_medida", "unidad de medida", "unidad", "Unidad de medida"],
        "precio_costo": ["precio_costo", "precio costo", "costo", "Precio costo"],
        "precio_venta": ["precio_venta", "precio venta", "precio", "Precio venta"],
        "precio_mayorista": ["precio_mayorista", "precio mayorista", "mayorista", "Precio mayorista"],
        "impuesto": ["impuesto", "iva", "impuesto_porcentaje", "porcentaje impuesto", "Impuesto"],
        "codigo_cabys": ["codigo_cabys", "cabys", "código cabys", "CABYS", "código_cabys"]
    }
    result = {}
    for key, aliases in mapping.items():
        found = False
        for alias in aliases:
            for r_key in row.keys():
                if r_key.lower().replace("_", " ").strip() == alias.lower().replace("_", " ").strip():
                    result[key] = row[r_key]
                    found = True
                    break
            if found:
                break
        if not found:
            result[key] = ""
    return result

@router.post("/import/preview")
def preview_import(
    payload: Dict[str, Any],
    session: Session = Depends(get_session)
):
    file_base64 = payload.get("file_base64")
    file_name = payload.get("file_name", "archivo.csv")
    if not file_base64:
        raise HTTPException(status_code=400, detail="Base64 file data is required")
    try:
        if "," in file_base64:
            file_base64 = file_base64.split(",")[1]
        content_bytes = base64.b64decode(file_base64)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Base64: {e}")
    is_xlsx = file_name.endswith(".xlsx")
    try:
        if is_xlsx:
            raw_rows = parse_xlsx(content_bytes)
        else:
            raw_rows = parse_csv(content_bytes)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error parsing file: {e}")
    preview_items = []
    seen_skus = set()
    seen_barcodes = set()
    db_products = session.exec(select(Producto)).all()
    db_skus = {p.sku for p in db_products}
    db_barcodes = {p.codigo_barras for p in db_products}
    has_errors = False
    for idx, raw_row in enumerate(raw_rows, start=1):
        mapped = map_fields(raw_row)
        errors = []
        warnings = []
        sku = mapped.get("sku", "").strip()
        barcode = mapped.get("codigo_barras", "").strip()
        nombre = mapped.get("nombre", "").strip()
        if not sku:
            errors.append("SKU es requerido")
        if not barcode:
            errors.append("Código de barras es requerido")
        if not nombre:
            errors.append("Nombre es requerido")
        if sku:
            if sku in seen_skus:
                errors.append(f"SKU '{sku}' duplicado en el archivo")
            else:
                seen_skus.add(sku)
        if barcode:
            if barcode in seen_barcodes:
                errors.append(f"Código de barras '{barcode}' duplicado en el archivo")
            else:
                seen_barcodes.add(barcode)
        if sku and sku in db_skus:
            errors.append(f"El SKU '{sku}' ya existe en el sistema")
        if barcode and barcode in db_barcodes:
            errors.append(f"El código de barras '{barcode}' ya existe en el sistema")
        try:
            costo = float(mapped.get("precio_costo") or 0)
            if costo < 0:
                errors.append("Precio costo no puede ser negativo")
        except ValueError:
            errors.append("Precio costo inválido")
            costo = 0.0
        try:
            venta = float(mapped.get("precio_venta") or 0)
            if venta < 0:
                errors.append("Precio venta no puede ser negativo")
        except ValueError:
            errors.append("Precio venta inválido")
            venta = 0.0
        try:
            mayorista = float(mapped.get("precio_mayorista") or 0)
            if mayorista < 0:
                errors.append("Precio mayorista no puede ser negativo")
        except ValueError:
            errors.append("Precio mayorista inválido")
            mayorista = 0.0
        if len(errors) > 0:
            has_errors = True
        categoria = mapped.get("categoria", "").strip()
        subcategoria = mapped.get("subcategoria", "").strip()
        marca = mapped.get("marca", "").strip()
        if not categoria:
            warnings.append("Categoría vacía. Se usará 'General'")
        if not subcategoria:
            warnings.append("Subcategoría vacía. Se usará 'General'")
        if not marca:
            warnings.append("Marca vacía. Se usará 'General'")
        preview_items.append({
            "index": idx,
            "data": {
                "sku": sku,
                "codigo_barras": barcode,
                "nombre": nombre,
                "descripcion": mapped.get("descripcion", ""),
                "categoria": categoria or "General",
                "subcategoria": subcategoria or "General",
                "marca": marca or "General",
                "unidad_medida": mapped.get("unidad_medida") or "Unidad",
                "precio_costo": costo,
                "precio_venta": venta,
                "precio_mayorista": mayorista,
                "impuesto": mapped.get("impuesto") or "IVA 13%"
            },
            "errors": errors,
            "warnings": warnings
        })
    return {
        "filename": file_name,
        "total_rows": len(raw_rows),
        "has_errors": has_errors,
        "items": preview_items
    }

@router.post("/import/confirm")
def confirm_import(
    payload: Dict[str, Any],
    request: Request,
    current_user: Usuario = Depends(PermissionChecker("inventory:edit")),
    session: Session = Depends(get_session)
):
    items = payload.get("items", [])
    if not items:
        raise HTTPException(status_code=400, detail="No items to import")
    count = 0
    default_tax = session.exec(select(Impuesto).where(Impuesto.porcentaje == 13.0)).first()
    if not default_tax:
        default_tax = session.exec(select(Impuesto)).first()
    for item in items:
        data = item.get("data", {})
        sku = data.get("sku")
        barcode = data.get("codigo_barras")
        nombre = data.get("nombre")
        if not sku or not barcode or not nombre:
            continue
        dup = session.exec(select(Producto).where(or_(Producto.sku == sku, Producto.codigo_barras == barcode))).first()
        if dup:
            continue
        brand_name = data.get("marca", "General").strip()
        brand = session.exec(select(Marca).where(Marca.nombre == brand_name)).first()
        if not brand:
            brand = Marca(nombre=brand_name)
            session.add(brand)
            session.commit()
            session.refresh(brand)
        cat_name = data.get("categoria", "General").strip()
        cat = session.exec(select(Categoria).where(Categoria.nombre == cat_name)).first()
        if not cat:
            cat = Categoria(nombre=cat_name, dias_alerta_vencimiento=30)
            session.add(cat)
            session.commit()
            session.refresh(cat)
        subcat_name = data.get("subcategoria", "General").strip()
        subcat = session.exec(select(Subcategoria).where(
            Subcategoria.nombre == subcat_name,
            Subcategoria.categoria_id == cat.id
        )).first()
        if not subcat:
            subcat = Subcategoria(nombre=subcat_name, categoria_id=cat.id)
            session.add(subcat)
            session.commit()
            session.refresh(subcat)
        tax_str = data.get("impuesto", "IVA 13%").strip()
        tax = session.exec(select(Impuesto).where(Impuesto.nombre == tax_str)).first()
        if not tax:
            try:
                pct_str = ''.join(c for c in tax_str if c.isdigit() or c == '.')
                if pct_str:
                    pct = float(pct_str)
                    tax = session.exec(select(Impuesto).where(Impuesto.porcentaje == pct)).first()
            except:
                pass
        if not tax:
            tax = default_tax
        prod = Producto(
            sku=sku,
            codigo_barras=barcode,
            nombre=nombre,
            descripcion=data.get("descripcion", ""),
            marca_id=brand.id,
            subcategoria_id=subcat.id,
            unidad_medida=data.get("unidad_medida", "Unidad"),
            precio_costo=float(data.get("precio_costo") or 0.0),
            precio_venta=float(data.get("precio_venta") or 0.0),
            precio_mayorista=float(data.get("precio_mayorista") or 0.0),
            impuesto_id=tax.id if tax else 1,
            stock_minimo=0.0,
            stock_maximo=100.0,
            codigo_cabys=data.get("codigo_cabys") or None,
            activo=True
        )
        session.add(prod)
        session.commit()
        session.refresh(prod)
        inv = Inventario(
            sucursal_id=current_user.sucursal_id,
            producto_id=prod.id,
            existencia_actual=0.0
        )
        session.add(inv)
        session.commit()
        count += 1
    audit = Auditoria(
        usuario_id=current_user.id,
        accion="IMPORTACION_PRODUCTOS",
        modulo="Inventario",
        detalles=f"Importación masiva confirmada. Creados {count} productos con éxito.",
        ip_address=request.client.host
    )
    session.add(audit)
    session.commit()
    return {"message": f"Se han importado {count} productos con éxito."}

@router.get("/export")
def export_inventory(
    format: str = Query("csv", description="Formato de exportación: csv, excel, pdf"),
    filter_type: str = Query("all", description="Filtro de stock: all, stock_bajo, inactivos"),
    current_user: Usuario = Depends(get_current_user),
    session: Session = Depends(get_session)
):
    empresa = session.exec(select(Empresa)).first()
    if not empresa:
        empresa = Empresa(
            nombre_comercial="Minisúper M Y M",
            razon_social="Minisúper M Y M S.A.",
            cedula_juridica="3-101-000000",
            direccion="Costa Rica",
            telefonos="0000-0000",
            correo="contacto@minisupermym.com"
        )
    query = select(Producto)
    if filter_type == "inactivos":
        query = query.where(Producto.activo == False)
    else:
        query = query.where(Producto.activo == True)
    products = session.exec(query).all()
    items_data = []
    for p in products:
        inv = session.exec(select(Inventario).where(
            Inventario.producto_id == p.id,
            Inventario.sucursal_id == current_user.sucursal_id
        )).first()
        existencia = inv.existencia_actual if inv else 0.0
        if filter_type == "stock_bajo" and existencia > p.stock_minimo:
            continue
        items_data.append({
            "sku": p.sku,
            "codigo_barras": p.codigo_barras,
            "nombre": p.nombre,
            "marca": p.marca.nombre if p.marca else "N/A",
            "existencia": existencia,
            "precio_costo": p.precio_costo,
            "precio_venta": p.precio_venta,
            "activo": p.activo
        })
    if format == "pdf":
        from fastapi.responses import FileResponse
        import tempfile
        temp_dir = tempfile.gettempdir()
        filename = os.path.join(temp_dir, f"reporte_inventario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")
        from ..utils.pdf_generator import generate_inventory_pdf
        generate_inventory_pdf(empresa, items_data, filter_type, filename)
        audit = Auditoria(
            usuario_id=current_user.id,
            accion="EXPORTACION_INVENTARIO",
            modulo="Inventario",
            detalles=f"Exportación de inventario en PDF. Registros: {len(items_data)}",
            ip_address="127.0.0.1"
        )
        session.add(audit)
        session.commit()
        return FileResponse(filename, media_type="application/pdf", filename="reporte_inventario.pdf")
    elif format == "excel" or format == "xlsx":
        from fastapi.responses import StreamingResponse
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventario"
        headers = ["SKU", "Código de barras", "Nombre", "Marca", "Existencia", "Costo Unit.", "Precio Venta", "Estado"]
        ws.append(headers)
        for item in items_data:
            ws.append([
                item["sku"],
                item["codigo_barras"],
                item["nombre"],
                item["marca"],
                item["existencia"],
                item["precio_costo"],
                item["precio_venta"],
                "Activo" if item["activo"] else "Inactivo"
            ])
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        audit = Auditoria(
            usuario_id=current_user.id,
            accion="EXPORTACION_INVENTARIO",
            modulo="Inventario",
            detalles=f"Exportación de inventario en Excel. Registros: {len(items_data)}",
            ip_address="127.0.0.1"
        )
        session.add(audit)
        session.commit()
        return StreamingResponse(
            out, 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=reporte_inventario.xlsx"}
        )
    else: # csv
        from fastapi.responses import StreamingResponse
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["SKU", "Codigo Barras", "Nombre", "Marca", "Existencia", "Precio Costo", "Precio Venta", "Estado"])
        for item in items_data:
            writer.writerow([
                item["sku"],
                item["codigo_barras"],
                item["nombre"],
                item["marca"],
                item["existencia"],
                item["precio_costo"],
                item["precio_venta"],
                "Activo" if item["activo"] else "Inactivo"
            ])
        output.seek(0)
        content = output.getvalue()
        audit = Auditoria(
            usuario_id=current_user.id,
            accion="EXPORTACION_INVENTARIO",
            modulo="Inventario",
            detalles=f"Exportación de inventario en CSV. Registros: {len(items_data)}",
            ip_address="127.0.0.1"
        )
        session.add(audit)
        session.commit()
        return StreamingResponse(
            io.BytesIO(content.encode("utf-8")), 
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=reporte_inventario.csv"}
        )

@router.put("/lots/{id}/expiry")
def update_lot_expiry(
    id: int,
    payload: Dict[str, Any],
    request: Request,
    current_user: Usuario = Depends(PermissionChecker("inventory:edit")),
    session: Session = Depends(get_session)
):
    lote = session.get(Lote, id)
    if not lote:
        raise HTTPException(status_code=404, detail="Lote no encontrado")
    fecha_str = payload.get("fecha_vencimiento")
    motivo = payload.get("motivo")
    if not fecha_str or not motivo:
        raise HTTPException(status_code=400, detail="La fecha de vencimiento y el motivo son obligatorios")
    try:
        nueva_fecha = datetime.strptime(fecha_str, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido. Usar YYYY-MM-DD")
    fecha_anterior = lote.fecha_vencimiento
    lote.fecha_vencimiento = nueva_fecha
    session.add(lote)
    audit = Auditoria(
        usuario_id=current_user.id,
        accion="LOT_EXPIRY_UPDATE",
        modulo="Inventario",
        detalles=f"Vencimiento de lote {lote.numero_lote} (Producto: {lote.producto.nombre}) modificado de {fecha_anterior} a {nueva_fecha}. Motivo: {motivo}",
        ip_address=request.client.host
    )
    session.add(audit)
    session.commit()
    return {"message": "Fecha de vencimiento del lote actualizada con éxito"}

# --- CATÁLOGO CABYS ---

@router.get("/cabys", response_model=List[Dict[str, Any]])
def search_cabys(
    q: str = Query(..., description="Buscar por descripción o código de 13 dígitos"),
    session: Session = Depends(get_session)
):
    from ..database.schema import Cabys
    query_str = q.strip()
    if not query_str:
        return []
        
    if query_str.isdigit():
        stmt = select(Cabys).where(Cabys.codigo.like(f"%{query_str}%")).limit(50)
    else:
        words = [w.strip() for w in query_str.split(" ") if w.strip()]
        stmt = select(Cabys)
        for w in words:
            stmt = stmt.where(Cabys.descripcion.like(f"%{w}%"))
        stmt = stmt.limit(50)
        
    results = session.exec(stmt).all()
    return [
        {
            "codigo": c.codigo,
            "descripcion": c.descripcion,
            "impuesto": c.impuesto
        }
        for c in results
    ]

@router.get("/cabys/status")
def get_cabys_status(
    session: Session = Depends(get_session)
):
    from ..database.schema import Configuracion
    
    status = session.exec(select(Configuracion).where(Configuracion.clave == "cabys_sync_status")).first()
    last_update = session.exec(select(Configuracion).where(Configuracion.clave == "cabys_last_update")).first()
    total_records = session.exec(select(Configuracion).where(Configuracion.clave == "cabys_total_records")).first()
    error = session.exec(select(Configuracion).where(Configuracion.clave == "cabys_sync_error")).first()
    
    return {
        "status": status.valor if status else "idle",
        "last_update": last_update.valor if last_update else None,
        "total_records": int(total_records.valor) if total_records and total_records.valor.isdigit() else 0,
        "error": error.valor if error else ""
    }

@router.post("/cabys/sync")
def trigger_cabys_sync(
    current_user: Usuario = Depends(PermissionChecker("settings:edit")),
    session: Session = Depends(get_session)
):
    from ..utils.cabys_sync import run_sync_in_background, get_config_val
    
    status = get_config_val(session, "cabys_sync_status", "idle")
    if status == "syncing":
        raise HTTPException(status_code=400, detail="Ya se encuentra en proceso una sincronización del catálogo.")
        
    # Lanzar la sincronización en segundo plano de manera no bloqueante
    run_sync_in_background()
    return {"message": "Sincronización del catálogo CABYS iniciada en segundo plano."}


